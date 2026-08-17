# Standard library
import io
import math
import re
from datetime import datetime
from pathlib import Path

# Third-party
import streamlit as st
from openpyxl import load_workbook

#python -m streamlit run .\streamlit_app.py

# Optional pandas (used for table processing on Conduit page)
try:
    import pandas as pd  # type: ignore
except ImportError:
    pd = None  # type: ignore

# Local application imports
try:
    from lib.theory import render_md, render_md_text  # type: ignore
    _THEORY_IMPORT_ERROR = None
except Exception as e:
    render_md = None  # type: ignore
    render_md_text = None  # type: ignore
    _THEORY_IMPORT_ERROR = str(e)

try:
    from lib import oesc_tables  # type: ignore
    _TABLES_IMPORT_ERROR = None
except Exception as e:
    oesc_tables = None  # type: ignore
    _TABLES_IMPORT_ERROR = str(e)

# Set to False during development to disable password protection
ENABLE_PASSWORD_PROTECTION = True

# ----------------------------
# Global Variables
# ----------------------------
PROJECT_NUMBER = ""
DESIGNER_NAME = ""
PANEL_TEMPLATE_PATH = Path("content/files/panel_schedule_template.xlsx")
SUPPORTED_CODE_VERSIONS = {
    "NEC": ["2026"],
    "OESC": ["2024"]
}

# ----------------------------
# Data / math utilities
# ----------------------------










# ----------------------------
# UI utilities
# ----------------------------


# ----------------------------
# Word document helpers
# ----------------------------


















# ----------------------------
# Excel helpers
# ----------------------------






# ----------------------------
# Export button renderer
# ----------------------------


# ----------------------------
# Geometry utilities
# ----------------------------




st.set_page_config(
    page_title="Electrical Calculations Hub",
    page_icon="⚡",
    layout="wide",
)


# Center all images (both st.image and markdown-rendered images)
st.markdown(
    """
<style>
img { display: block; margin-left: auto; margin-right: auto; }
.stImage { text-align: center; }
/* Hide horizontal scrollbars on LaTeX blocks (keeps content scrollable if needed) */
.katex-display { scrollbar-width: none; -ms-overflow-style: none; }
.katex-display::-webkit-scrollbar { display: none; }
/* Compact expanders fade and shrink their label; match the lists around them */
[class*="st-key-appendix-tables"] .stExpander summary {
  opacity: 1 !important;
}
[class*="st-key-appendix-tables"] .stExpander summary [data-testid="stMarkdownContainer"],
[class*="st-key-appendix-tables"] .stExpander summary [data-testid="stMarkdownContainer"] p {
  font-size: 1rem !important;
  color: inherit !important;
}
</style>
""",
    unsafe_allow_html=True,
)

# ----------------------------
# Password Protection
# ----------------------------
def check_password():
    """Returns `True` if the user had the correct password."""

    def password_entered():
        """Checks whether a password entered by the user is correct."""
        try:
            admin_password = st.secrets.get("app_password_admin", "admin")
            user_password = st.secrets.get("app_password_user", "JNE")
            legacy_password = st.secrets.get("app_password", None)
        except (KeyError, FileNotFoundError):
            admin_password = "admin"
            user_password = "JNE"
            legacy_password = None

        entered = st.session_state.get("password", "")
        if entered == admin_password or (legacy_password and entered == legacy_password):
            st.session_state["password_correct"] = True
            st.session_state["access_role"] = "admin"
            del st.session_state["password"]  # don't store password
        elif entered == user_password:
            st.session_state["password_correct"] = True
            st.session_state["access_role"] = "user"
            del st.session_state["password"]  # don't store password
        else:
            st.session_state["password_correct"] = False
            st.session_state["access_role"] = None

    if st.session_state.get("password_correct", False):
        return True

    st.text_input(
        "Enter password to access the site:",
        type="password",
        on_change=password_entered,
        key="password",
    )

    if "password_correct" in st.session_state and not st.session_state["password_correct"]:
        st.error("❌ Incorrect password")

    return False


if ENABLE_PASSWORD_PROTECTION and not check_password():
    st.stop()

st.title("⚡ Electrical Calculations Hub")
st.caption("Theory • Examples • Calculators")

# ----------------------------
# Formatting utilities
# ----------------------------
def fmt(x, unit=""):
    if x is None:
        return "—"
    try:
        x = float(x)
    except Exception:
        return str(x)
    if abs(x) >= 1e6:
        s = f"{x:,.3g}"
    elif abs(x) >= 1:
        s = f"{x:,.4g}"
    else:
        s = f"{x:.6g}"
    return f"{s} {unit}".strip()










# ----------------------------
# Calculation utilities
# ----------------------------

# Practical "standard" list used by the attached OESC calc (Table 13 style). This list is commonly aligned with the NEC list.








# ----------------------------
# UI helpers
# ----------------------------
def header(title: str, subtitle: str = ""):
    st.header(title)
    if subtitle:
        st.write(subtitle)


def show_code_note(selected_code: str):
    st.info(
        f"Code mode: **{selected_code} ({code_version})** "
        "This site is written to be easy to follow. Always verify final selections against the code, "
        "project specs, equipment data, and a coordination study where required."
    )


def eq(latex: str):
    """Render a LaTeX equation in a consistent display style."""
    st.latex(latex)


APP_DIR = Path(__file__).parent
CONTENT_DIR = APP_DIR / "content"


APPENDIX_SECTION_ORDER = (
    "Related Knowledge Files",
    "Related NEC Articles",
    "Related OESC Rules",
    "Related NEC Tables",
    "Related OESC Tables",
)

APPENDIX_TABLE_SECTIONS = {
    "Related NEC Tables": "NEC",
    "Related OESC Tables": "OESC",
}

_HTML_COMMENT_RE = re.compile(r"<!--.*?-->", re.S)
_APPENDIX_HEADING_RE = re.compile(r"^##\s+Appendix\s*$", re.M)
_APPENDIX_SECTION_RE = re.compile(r"^###\s+(.+?)\s*$(.*?)(?=^###\s|\Z)", re.M | re.S)
_APPENDIX_ENTRY_SPLIT_RE = re.compile(r"<br\s*/?>|\n")
_APPENDIX_REFERENCE_PREFIX_RE = re.compile(r"^Tables?\s+")


def split_appendix(md: str) -> tuple[str, dict[str, str]]:
    # Split a theory/examples document into its body and its appendix sections.
    
    match = _APPENDIX_HEADING_RE.search(md)
    if match is None:
        return md, {}

    body = md[: match.start()].rstrip()
    while body.endswith("---"):
        body = body[: -len("---")].rstrip()

    tail = _HTML_COMMENT_RE.sub("", md[match.end():])
    sections = {}
    for heading, content in _APPENDIX_SECTION_RE.findall(tail):
        content = content.strip()
        if content:
            sections[heading] = content
    return body, sections


def appendix_entries(content: str) -> list[tuple[str, str]]:
    entries = []
    for raw in _APPENDIX_ENTRY_SPLIT_RE.split(content):
        line = raw.strip()
        if line:
            reference, _, name = line.partition("—")
            entries.append((reference.strip(), name.strip()))
    return entries


def tables_module(code: str):
    if code == "OESC":
        return oesc_tables
    try:
        from lib import nec_tables  # type: ignore
        return nec_tables
    except Exception:
        return None


def _table_id(token: str, code: str) -> str:
    if code == "NEC":
        return "table_" + re.sub(r"[^a-z0-9]+", "_", token.lower()).strip("_")
    return re.sub(r"[^A-Z0-9]", "", token.upper())


def resolve_table_ids(reference: str, code: str, known_ids: set[str]) -> list[str]:
    # Handles "Table 1/2" and "Tables 9A-9G"
    body = _APPENDIX_REFERENCE_PREFIX_RE.sub("", reference.replace("Chapter 9,", "").strip())
    found = []
    for part in body.split("/"):
        start, dash, end = part.strip().partition("-")
        if dash:
            low, high = _table_id(start, code), _table_id(end, code)
            found += sorted(i for i in known_ids if low <= i <= high and len(i) == len(low))
            continue

        table_id = _table_id(start, code)
        if table_id in known_ids:
            found.append(table_id)
        elif code == "NEC":
            # 310.15(B)(1) is stored as table_310_15_b_1_1 and _1_2
            found += sorted(i for i in known_ids if i.startswith(f"{table_id}_"))
    return found


def _table_label(tables, table_id: str) -> str:
    meta = tables.get_table_meta(table_id) or {}
    return meta.get("title") or table_id.replace("table_", "Table ").replace("_", " ")


def render_table_entry(reference: str, name: str, code: str, tables, known_ids: set[str], scope: str):
    label = f"{reference} — {name}" if name else reference
    with st.expander(label, type="compact"):
        with st.container(border=True):
            table_ids = resolve_table_ids(reference, code, known_ids) if tables else []
            if not table_ids:
                st.caption("This table is not in the table library yet.")
                return

            for table_id in table_ids:
                meta = tables.get_table_meta(table_id) or {}

                # Parent tables hold no rows of their own
                parts = (meta.get("raw") or {}).get("tables") if meta.get("rows") is None else None
                if parts:
                    table_id = st.selectbox(
                        "Select a sub-table",
                        [f"{table_id}_{part}" for part in parts],
                        format_func=lambda tid: _table_label(tables, tid),
                        key=f"appendix_{scope}_{table_id}",
                    )
                    meta = tables.get_table_meta(table_id) or {}

                st.markdown(f"**{meta.get('title') or label}**")
                if meta.get("units"):
                    st.caption(f"Units: {meta['units']}")

                frame = tables.get_table_dataframe(table_id)
                if frame is None:
                    st.json(meta.get("raw", {}))
                else:
                    st.dataframe(frame, hide_index=True)
                
                st.divider()


def render_appendix(sections: dict[str, str], scope: str):
    ordered = [(t, sections[t]) for t in APPENDIX_SECTION_ORDER if t in sections]
    ordered += [(t, c) for t, c in sections.items() if t not in APPENDIX_SECTION_ORDER]
    if not ordered:
        return

    # Every element carries its own vertical gap, so batch the markdown
    pending = ["---", "## Appendix"]

    def flush():
        if pending:
            st.markdown("\n\n".join(pending), unsafe_allow_html=True)
            pending.clear()

    for title, content in ordered:
        pending.append(f"### {title}")
        code = APPENDIX_TABLE_SECTIONS.get(title)
        if code is None:
            pending.append(content)
            continue

        flush()
        tables = tables_module(code)
        known_ids = set(tables.search_tables("")) if tables else set()
        with st.container(gap=None, key=f"appendix-tables-{scope}"):
            for reference, name in appendix_entries(content):
                render_table_entry(reference, name, code, tables, known_ids, scope)

    flush()


def render_md_safe(rel_path: str):
    """
    Render markdown from /content safely.
    - Uses lib.theory.render_md if available
    - Otherwise shows a friendly error instead of crashing the app
    """
    md_path = CONTENT_DIR / rel_path

    if render_md is None or render_md_text is None:
        st.error(
            "Theory renderer failed to import. This usually means the `lib/` package is missing in your repo. "
            "Make sure you have:\n\n"
            "- `lib/__init__.py`\n"
            "- `lib/theory.py`\n"
        )
        if _THEORY_IMPORT_ERROR is not None:
            with st.expander("Import error details"):
                st.exception(_THEORY_IMPORT_ERROR)
        st.info(f"Expected markdown path: `{md_path}`")
        if md_path.exists():
            st.warning("Markdown file exists, but renderer is unavailable due to the import error above.")
        else:
            st.warning("Markdown file not found at the expected path.")
        return

    if not md_path.exists():
        st.error(f"Markdown file not found: {md_path}")
        return

    body, sections = split_appendix(md_path.read_text(encoding="utf-8"))
    render_md_text(body, md_path.parent)
    render_appendix(sections, md_path.stem)


def render_md_for_code(topic: str, code_mode: str):
    render_md_safe(f"markdown/{topic}_{'oesc' if code_mode == 'OESC' else 'nec'}.md")


# ----------------------------
# Panel schedule helpers
# ----------------------------
def _panel_safe_number(value):
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return value
    s = str(value).strip()
    if s == "":
        return None
    try:
        num = float(s)
    except Exception:
        return value
    return int(num) if abs(num - int(num)) < 1e-12 else num


def _panel_set_cell(ws, cell, value):
    ws[cell].value = None if value in ("", None) else value


def _panel_get_schedule_sheet(wb):
    for name in wb.sheetnames:
        if name.upper() != "COVER":
            return wb[name]
    return wb.active


def build_panel_schedule_workbook(
    cover_data,
    panel_header,
    left_rows,
    right_rows,
    use_current_date=False,
):
    wb = load_workbook(PANEL_TEMPLATE_PATH)
    cover = wb["COVER"]
    ws = _panel_get_schedule_sheet(wb)

    # ----------------------------
    # COVER sheet
    # ----------------------------
    _panel_set_cell(cover, "A1", cover_data.get("client"))
    _panel_set_cell(cover, "A2", cover_data.get("facility"))
    _panel_set_cell(cover, "A3", cover_data.get("building"))
    _panel_set_cell(cover, "A4", cover_data.get("title"))
    _panel_set_cell(cover, "A5", cover_data.get("panel_tag"))

    _panel_set_cell(cover, "F1", cover_data.get("drawn_by"))
    _panel_set_cell(cover, "F3", cover_data.get("checked_by"))
    _panel_set_cell(cover, "F5", cover_data.get("approved_by"))

    _panel_set_cell(cover, "G8", cover_data.get("project_number"))
    _panel_set_cell(cover, "I8", cover_data.get("drawing_number"))

    _panel_set_cell(cover, "G11", cover_data.get("cover_building"))
    _panel_set_cell(cover, "G12", cover_data.get("cover_title"))
    _panel_set_cell(cover, "A11", cover_data.get("notes"))

    _panel_set_cell(cover, "A15", cover_data.get("rev_no"))
    _panel_set_cell(cover, "B15", cover_data.get("revised_by"))
    _panel_set_cell(cover, "D15", cover_data.get("rev_checked_by"))
    _panel_set_cell(cover, "G15", cover_data.get("rev_approved_by"))
    _panel_set_cell(cover, "I15", cover_data.get("rev_notes"))

    if use_current_date:
        today = datetime.today().date()
        for cell in ["F2", "F4", "F6", "C15", "F15", "H15"]:
            cover[cell].value = today
    else:
        _panel_set_cell(cover, "F2", cover_data.get("drawn_date"))
        _panel_set_cell(cover, "F4", cover_data.get("checked_date"))
        _panel_set_cell(cover, "F6", cover_data.get("approved_date"))
        _panel_set_cell(cover, "C15", cover_data.get("rev_date"))
        _panel_set_cell(cover, "F15", cover_data.get("rev_checked_date"))
        _panel_set_cell(cover, "H15", cover_data.get("rev_approved_date"))

    # ----------------------------
    # PANEL schedule sheet header
    # ----------------------------
    _panel_set_cell(ws, "B1", panel_header.get("distribution_board_no"))
    _panel_set_cell(ws, "G1", panel_header.get("location"))
    _panel_set_cell(ws, "M1", panel_header.get("document_number"))
    _panel_set_cell(ws, "B2", panel_header.get("bus_rating"))
    _panel_set_cell(ws, "G2", panel_header.get("incomer_transformer_rating"))
    _panel_set_cell(ws, "M2", panel_header.get("equipment_layout_number"))
    _panel_set_cell(ws, "B3", panel_header.get("num_circuits"))
    _panel_set_cell(ws, "G3", panel_header.get("incomer_breaker"))
    _panel_set_cell(ws, "B4", panel_header.get("incomer_transformer_tag"))
    _panel_set_cell(ws, "G4", panel_header.get("interrupting_capacity"))

    # ----------------------------
    # Schedule rows (6-17)
    # ----------------------------
    for i in range(12):
        row = 6 + i
        left = left_rows[i] if i < len(left_rows) else {}
        right = right_rows[i] if i < len(right_rows) else {}

        _panel_set_cell(ws, f"A{row}", left.get("Load Description"))
        _panel_set_cell(ws, f"B{row}", _panel_safe_number(left.get("Conn Load (W)")))
        _panel_set_cell(ws, f"C{row}", left.get("RCCB Rating"))
        _panel_set_cell(ws, f"D{row}", _panel_safe_number(left.get("No. of Fixt.")))
        _panel_set_cell(ws, f"E{row}", left.get("Brkr Size"))
        _panel_set_cell(ws, f"F{row}", left.get("Cct No"))

        # Phase markers on left (G/H/I)
        ws[f"G{row}"].value = None
        ws[f"H{row}"].value = None
        ws[f"I{row}"].value = None
        phase = (left.get("Phase") or "").strip().upper()
        if phase == "A":
            ws[f"G{row}"].value = "A"
        elif phase == "B":
            ws[f"H{row}"].value = "B"
        elif phase == "C":
            ws[f"I{row}"].value = "C"

        _panel_set_cell(ws, f"J{row}", right.get("Cct No"))
        _panel_set_cell(ws, f"K{row}", right.get("Brkr Size"))
        _panel_set_cell(ws, f"L{row}", _panel_safe_number(right.get("No. of Fixt.")))
        _panel_set_cell(ws, f"M{row}", right.get("RCCB Rating"))
        _panel_set_cell(ws, f"N{row}", _panel_safe_number(right.get("Conn Load (W)")))
        _panel_set_cell(ws, f"O{row}", right.get("Load Description"))

    return wb


def default_panel_left_rows():
    return [
        {"Cct No": n, "Phase": p, "Load Description": "", "Conn Load (W)": "", "RCCB Rating": "", "No. of Fixt.": "", "Brkr Size": ""}
        for n, p in zip([1, 3, 5, 7, 9, 11, 13, 15, 17, 19, 21, 23], ["A", "B", "C", "A", "B", "C", "A", "B", "C", "A", "B", "C"])
    ]


def default_panel_right_rows():
    return [
        {"Cct No": n, "Brkr Size": "", "No. of Fixt.": "", "RCCB Rating": "", "Conn Load (W)": "", "Load Description": ""}
        for n in [2, 4, 6, 8, 10, 12, 14, 16, 18, 20, 22, 24]
    ]


# ----------------------------
# Sidebar navigation
# ----------------------------
ALL_PAGES = [
    "Home",
    "Conductors",
    "Voltage Drop",
    "Panel Schedule",
    "Transformer Feeders",
    "Transformer Protection",
    "Grounding/Bonding Conductor Sizing",
    "Motor Feeder",
    "Motor Protection",
    "Cable Tray Size & Fill & Bend Radius",
    "Conduit Size & Fill & Bend Radius",
    "Heat Trace",
    "Demand Load",
    "Power Factor Correction",
    "Table Library",
]

RESTRICTED_PAGES = {
    "Grounding/Bonding Conductor Sizing",
    "Demand Load",
    "Power Factor Correction",
    "Heat Trace",
    "Panel Schedule",
}

access_role = st.session_state.get("access_role", "admin")
if access_role == "user":
    PAGES = [p for p in ALL_PAGES if p not in RESTRICTED_PAGES]
else:
    PAGES = ALL_PAGES

with st.sidebar:
    st.header("Navigate")
    page = st.radio("Go to", PAGES, index=0)

    st.divider()
    st.header("Jurisdiction")
    
    available_codes = list(SUPPORTED_CODE_VERSIONS.keys())
    default_index = 1 if len(available_codes) > 1 and available_codes[1] == "OESC" else 0
    
    code_mode = st.selectbox("Select electrical code", available_codes, index=default_index)
    code_version = st.selectbox("Select version", SUPPORTED_CODE_VERSIONS.get(code_mode, []))

    st.divider()
    st.header("Report Information")
    PROJECT_NUMBER = st.text_input("Project number", value=PROJECT_NUMBER, key="project_number")
    DESIGNER_NAME = st.text_input("Designer name", value=DESIGNER_NAME, key="designer_name")    

    st.divider()
    with st.expander("🐛 Report an Issue / Request a Feature"):
        issue_type = st.selectbox(
            "Type",
            ["Bug Report", "Feature Request"],
            key="issue_type"
        )
        issue_text = st.text_area(
            "Describe the issue or request",
            key="issue_text",
            height=150,
            placeholder="Please describe the problem or feature as clearly as possible..."
        )
        if st.button("Send Report", key="submit_issue"):
            if issue_text.strip():
                import urllib.parse
                recipients = "kmurphy@jnegroup.com,DFritzsche@jnegroup.com,NZuvela@jnegroup.com"
                subject = urllib.parse.quote(f"[{issue_type}] JNE Calculations Portal")
                body = urllib.parse.quote(issue_text)
                st.link_button(
                    "📧 Click here to open in Outlook",
                    f"mailto:{recipients}?subject={subject}&body={body}",
                )
            else:
                st.warning("Please describe the issue before submitting.")

    st.caption("This portal is provided for educational purposes only and is intended to support the understanding of engineering concepts. The tutorials, examples, and tools are not a substitute for professional judgment. Always consult applicable codes, regulations, and qualified engineers before making design or compliance decisions.")

    st.markdown("---")
    st.caption("Rev. 1.0")


# ----------------------------
# Page shell with Theory/Examples/Calculator tabs
# (Tabs are disabled ONLY on Table Library)
# ----------------------------
if page not in ("Table Library", "Home"):
    theory_tab, examples_tab, calc_tab = st.tabs(["📚 Theory", "🧩 Examples", "🧮 Calculator"])
else:
    theory_tab = None
    examples_tab = None
    calc_tab = None


# ============================
# 0) Home
# ============================
if page == "Home":
    header("Welcome", "Start here for quick context and how to use this hub.")
    show_code_note(code_mode)

    st.markdown("### What you can do")
    st.markdown("- Find code-aligned theory notes with worked examples.")
    st.markdown("- Run calculators for sizing, protection, and voltage drop.")
    st.markdown("- Compare NEC vs OESC assumptions using the sidebar selector.")

    st.markdown("### Popular tools")
    st.markdown("- Transformer Protection")
    st.markdown("- Voltage Drop")
    st.markdown("- Conduit Size & Fill & Bend Radius")
    st.markdown("- Cable Tray Size & Fill & Bend Radius")

    st.markdown("### Quick start")
    st.markdown("1. Pick a topic from the sidebar.")
    st.markdown("2. Use the `Theory` tab for context and code references.")
    st.markdown("3. Review the `Examples` tab for worked examples.")
    st.markdown("4. Switch to `Calculator` for inputs and results.")
    st.markdown("5. Change `Jurisdiction` to see NEC vs OESC logic.")


# ============================
# 1) Transformer Protection
# ============================
if page == "Transformer Protection":
    with theory_tab:
        header("Transformer Protection")
        show_code_note(code_mode)

        render_md_for_code("transformer_protection", code_mode)

    with examples_tab:
        header("Transformer Protection — Examples")
        show_code_note(code_mode)
        render_md_for_code("transformer_protection_examples", code_mode)

    # ----------------------------
    # Calculator tab for Transformer Protection
    # ----------------------------
    with calc_tab:
        if code_mode == "NEC":
            header("Transformer Protection — Calculator", "NEC Article 450.3 primary and secondary overcurrent sizing.")
            show_code_note(code_mode)
            from nec_calc.transformer_protection.ui import render_calc as render_nec_tp
            render_nec_tp()
        else:
            header("Transformer Protection Calculator", "OESC Rules 26-250, 26-252 and 26-254 overcurrent sizing.")
            show_code_note(code_mode)
            from oesc_calc.transformer_protection.ui import render_calc as render_oesc_tp
            render_oesc_tp()


# ============================
# 2) Transformer Feeders
# ============================
elif page == "Transformer Feeders":
    with theory_tab:
        header("Transformer Feeders — Theory")
        show_code_note(code_mode)
        render_md_for_code("transformer_feeders", code_mode)

    with examples_tab:
        header("Transformer Feeders — Examples")
        show_code_note(code_mode)
        render_md_for_code("transformer_feeders_examples", code_mode)

    with calc_tab:
        if code_mode == "NEC":
            header("Transformer Feeders — Calculator", "NEC Article 450.3 & 310 full-load current calculation.")
            show_code_note(code_mode)
            from nec_calc.transformer_feeder.ui import render_calc as render_nec_tf
            render_nec_tf()
        else:
            header("Transformer Feeder Calculator", "OESC Section 26 full-load current calculation.")
            show_code_note(code_mode)
            from oesc_calc.transformer_feeder.ui import render_calc as render_oesc_tf
            render_oesc_tf()


# ============================
# 3) Grounding/Bonding Conductor Sizing
# ============================
elif page == "Grounding/Bonding Conductor Sizing":
    with theory_tab:
        header("Grounding & Bonding — Theory")
        show_code_note(code_mode)
        render_md_for_code("grounding_bonding", code_mode)

    with examples_tab:
        header("Grounding & Bonding — Examples")
        show_code_note(code_mode)
        render_md_for_code("grounding_bonding_examples", code_mode)

    with calc_tab:
        header("Grounding/Bonding Helper", "Simple placeholder — replace with real NEC/OESC table logic.")
        show_code_note(code_mode)

        ocpd = st.number_input("Upstream OCPD rating (A)", min_value=1.0, value=200.0, step=1.0)

        if ocpd <= 60:
            egc = "10 AWG Cu (placeholder)"
        elif ocpd <= 100:
            egc = "8 AWG Cu (placeholder)"
        elif ocpd <= 200:
            egc = "6 AWG Cu (placeholder)"
        elif ocpd <= 400:
            egc = "3 AWG Cu (placeholder)"
        else:
            egc = "See table / engineer (placeholder)"

        st.success(f"Equipment grounding conductor (example placeholder): **{egc}**")
        st.markdown("### Equation used")
        eq(r"\text{EGC size} = f(\text{OCPD rating})")


# ============================
# 4) Motor Protection
# ============================
elif page == "Motor Protection":
    with theory_tab:
        header("Motor Protection — Theory")
        show_code_note(code_mode)
        render_md_for_code("motor_protection", code_mode)

    with examples_tab:
        header("Motor Protection — Examples")
        show_code_note(code_mode)
        render_md_for_code("motor_protection_examples", code_mode)

    with calc_tab:
      if code_mode == "NEC":
        header("Motor Protection — Calculator", "NEC Article 430 branch-circuit, overload, feeder and disconnect sizing.")
        show_code_note(code_mode)
        from nec_calc.motor_protection.ui import render_calc as render_nec_mp
        render_nec_mp()
      else:
        header("Motor Protection Calculator", "OESC Table 29 motor branch-circuit overcurrent device sizing.")
        show_code_note(code_mode)
        from oesc_calc.motor_protection.ui import render_calc as render_oesc_mp
        render_oesc_mp()


# ============================
# 5) Motor Feeder
# ============================
elif page == "Motor Feeder":
    with theory_tab:
        header("Motor Feeder — Theory")
        show_code_note(code_mode)
        render_md_for_code("motor_feeder", code_mode)

    with examples_tab:
        header("Motor Feeder — Examples")
        show_code_note(code_mode)
        render_md_for_code("motor_feeder_examples", code_mode)

    with calc_tab:
        if code_mode == "NEC":
            header("Motor Feeder — Calculator", "Single-motor feeder conductor (NEC 430.22) and overload (NEC 430.32) sizing, with full-load current per NEC 430.6(A).")
            show_code_note(code_mode)
            from nec_calc.motor_feeder.ui import render_calc as render_nec_mf
            render_nec_mf()
        else:
            header("Motor Feeder Calculator", "OESC Rule 28-106 motor feeder conductor sizing from nameplate data.")
            show_code_note(code_mode)
            from oesc_calc.motor_feeder.ui import render_calc as render_oesc_mf
            render_oesc_mf()


# ============================
# 6) Cable Tray Size & Fill & Bend Radius
# ============================
elif page == "Cable Tray Size & Fill & Bend Radius":
    with theory_tab:
        header("Cable Tray Size, Fill & Bend Radius — Theory")
        show_code_note(code_mode)
        render_md_for_code("cable_tray_fill", code_mode)

    with examples_tab:
        header("Cable Tray Size, Fill & Bend Radius — Examples")
        show_code_note(code_mode)
        render_md_for_code("cable_tray_fill_examples", code_mode)

    with calc_tab:
        if code_mode == "NEC":
            header("Cable Tray Fill Calculator — NEC", "NEC Article 392 cable tray fill requirements.")
            show_code_note(code_mode)
            from nec_calc.cable_tray_fill.ui import render_calc as render_nec_ctf
            render_nec_ctf()
        else:
            header("Cable Tray Fill Calculator", "Tray cross-section fill from cable diameters and quantities.")
            show_code_note(code_mode)
            from oesc_calc.cable_tray_fill.ui import render_calc as render_oesc_ctf
            render_oesc_ctf()


    # ============================
    # 7) Conduit Size & Fill & Bend Radius
    # ============================
elif page == "Conduit Size & Fill & Bend Radius":
    with theory_tab:
        header("Conduit Size, Fill & Bend Radius — Theory")
        show_code_note(code_mode)
        render_md_for_code("conduit_fill", code_mode)

    with examples_tab:
        header("Conduit Size, Fill & Bend Radius — Examples")
        show_code_note(code_mode)
        render_md_for_code("conduit_fill_examples", code_mode)

    with calc_tab:
        if code_mode == "NEC":
            header("Conduit Size, Fill & Bend Radius — Calculator", "NEC Chapter 9 conduit fill (Tables 1, 4, 5), minimum size, and bend radius (Table 2).")
            show_code_note(code_mode)
            from nec_calc.conduit_fill.ui import render_calc as render_nec_cf
            render_nec_cf()
        else:
            header("Conduit Fill Calculator", "OESC Tables 9A to 9H conduit fill and minimum trade size.")
            show_code_note(code_mode)
            from oesc_calc.conduit_fill.ui import render_calc as render_oesc_cf
            render_oesc_cf()


# ============================
# 8) Heat Trace
# ============================
elif page == "Heat Trace":
    with theory_tab:
        header("Heat Trace — Theory")
        show_code_note(code_mode)
        render_md_for_code("heat_trace", code_mode)

    with examples_tab:
        header("Heat Trace — Examples")
        show_code_note(code_mode)
        render_md_for_code("heat_trace_examples", code_mode)

    with calc_tab:
        header("Heat Trace Load Calculator", "Estimate circuit load from length and heat trace rating.")
        show_code_note(code_mode)

        length_m = st.number_input("Heat trace length (m)", min_value=0.1, value=60.0, step=1.0)
        watts_per_m = st.number_input("Heat output (W/m)", min_value=1.0, value=30.0, step=1.0)
        voltage = st.number_input("Supply voltage (V)", min_value=12.0, value=120.0, step=1.0)

        total_power_w = length_m * watts_per_m
        current_a = total_power_w / voltage if voltage > 0 else None
        st.success(f"Total heat trace load: **{fmt(total_power_w, 'W')}**")
        st.metric("Estimated circuit current", fmt(current_a, "A"))

        st.markdown("### Equation used")
        eq(r"P_{total}=L\cdot q,\ \ I=P_{total}/V")


# ============================
# 9) Demand Load
# ============================
elif page == "Demand Load":
    with theory_tab:
        header("Demand Load — Theory")
        show_code_note(code_mode)
        render_md_for_code("demand_load", code_mode)

    with examples_tab:
        header("Demand Load — Examples")
        show_code_note(code_mode)
        render_md_for_code("demand_load_examples", code_mode)

    with calc_tab:
        header("Demand Load Calculator", "Compute demand load from connected load and factor.")
        show_code_note(code_mode)

        connected = st.number_input("Connected load (kW)", min_value=0.0, value=120.0, step=1.0)
        factor = st.number_input("Demand factor (0–1)", min_value=0.0, max_value=1.0, value=0.65, step=0.01)
        demand = connected * factor
        st.success(f"Demand load: **{fmt(demand, 'kW')}**")

        st.markdown("### Equation used")
        eq(r"P_{demand}=P_{connected}\cdot f_{demand}")


# ============================
# 10) Power Factor Correction
# ============================
elif page == "Power Factor Correction":
    with theory_tab:
        header("Power Factor Correction — Theory")
        show_code_note(code_mode)
        render_md_for_code("power_factor_correction", code_mode)

    with examples_tab:
        header("Power Factor Correction — Examples")
        show_code_note(code_mode)
        render_md_for_code("power_factor_correction_examples", code_mode)

    with calc_tab:
        header("Power Factor Correction — Calculator")
        show_code_note(code_mode)
        st.info("Placeholder — content coming soon.")


# ============================
# Table Library (browse/search embedded OESC tables)
# ============================
elif page == "Table Library":

    header(f"Table Library — {code_mode} Tables")
    show_code_note(code_mode)

    tl_tab, qr_tab = st.tabs(["📚 Table Library", "📋 Quick Reference"])

    with tl_tab:
        if code_mode == "NEC":
            try:
                from lib import nec_tables
                _NEC_TABLES_ERROR = None
            except Exception as e:
                nec_tables = None
                _NEC_TABLES_ERROR = str(e)
            
            active_tables = nec_tables
            active_error = _NEC_TABLES_ERROR
        else:
            active_tables = oesc_tables
            active_error = _TABLES_IMPORT_ERROR

        if active_error:
            st.error(f"Table library failed to import: `{active_error}`")
        else:
            q = st.text_input(
                "Search tables",
                value="",
                placeholder="Examples: 1, 2, 5A, 9H, ampacity, conduit fill …",
            )

            table_ids = active_tables.search_tables(q)

            if not table_ids:
                st.warning("No tables match your search.")
            else:
                def _label(tid: str) -> str:
                    meta = active_tables.get_table_meta(tid)
                    title = meta.get("title") if meta else ""
                    
                    if title and title.lower().startswith("table"):
                        return title
                        
                    clean_tid = tid.replace("table_", "").replace("_", " ").upper()
                    return f"Table {clean_tid} — {title}" if title else f"Table {clean_tid}"

                selected = st.selectbox("Select a table", table_ids, format_func=_label)
                meta = active_tables.get_table_meta(selected) or {}
                
                # If this is a parent table, offer a sub-table selector
                if meta.get("rows") is None and "tables" in meta.get("raw", {}):
                    sub_keys = list(meta["raw"]["tables"].keys())
                    sub_ids = [f"{selected}_{k}" for k in sub_keys]
                    
                    sub_selected = st.selectbox("↳ Select specific sub-table", sub_ids, format_func=_label)
                    selected = sub_selected
                    meta = active_tables.get_table_meta(selected) or {}

                st.markdown(f"### {_label(selected)}")

                if meta.get("units"):
                    st.caption(f"Units: **{meta['units']}**")
                if meta.get("condition"):
                    st.caption(f"Applies to: **{meta['condition']}**")
                if meta.get("edition") or meta.get("source"):
                    _prov = " · ".join(
                        str(meta[k]) for k in ("edition", "source") if meta.get(k)
                    )
                    st.caption(f"Source: {_prov}")
                for _note in meta.get("notes") or []:
                    st.caption(_note)

                _DIAG_IMAGES = {
                    "D8":  ([CONTENT_DIR / "images" / "diagram_d8.png"], 650),
                    "D9":  ([CONTENT_DIR / "images" / "diagram_d9.png"], 950),
                    "D10": ([CONTENT_DIR / "images" / "diagram_d10a.png",
                             CONTENT_DIR / "images" / "diagram_d10b.png"], 650),
                    "D11": ([CONTENT_DIR / "images" / "diagram_d11a.png",
                             CONTENT_DIR / "images" / "diagram_d11b.png"], 650),
                }
                _sel_upper = str(selected).upper()
                for _diag_key, (_diag_imgs, _diag_w) in _DIAG_IMAGES.items():
                    if _sel_upper.startswith(_diag_key):
                        with st.expander(f"Diagram {_diag_key} — installation configurations", expanded=True):
                            if len(_diag_imgs) == 1:
                                if _diag_imgs[0].exists():
                                    _c1, _c2, _c3 = st.columns([1, 2, 1])
                                    _c2.image(str(_diag_imgs[0]), width=_diag_w)
                            else:
                                _cols = st.columns(len(_diag_imgs))
                                for _col, _img in zip(_cols, _diag_imgs):
                                    if _img.exists():
                                        _col.image(str(_img), width=650)
                        break

                df = active_tables.get_table_dataframe(selected)

                if df is None:
                    st.info("This table is stored in raw format.")
                    st.json(meta.get("raw", {}))
                else:
                    # Where the codebook prints a banded header, show the bands: stack the
                    # tiers over each column so the table reads like the printed page.
                    if meta.get("header_tiers") is not None and pd is not None:
                        _tiers = meta["header_tiers"]
                        _cols = meta.get("columns") or []
                        
                        # Resolve tiers horizontally (banded headers continue to the right)
                        _resolved_tiers = []
                        for _tier in _tiers:
                            _resolved_tier = []
                            _current = ""
                            for _i in range(len(_cols)):
                                _cell = _tier[_i].strip() if _i < len(_tier) else ""
                                if _cell:
                                    _current = _cell
                                _resolved_tier.append(_current)
                            _resolved_tiers.append(_resolved_tier)
                            
                        # Stack them vertically for Pandas MultiIndex
                        _stacked = []
                        for _i in range(len(_cols)):
                            _path = tuple(_rt[_i] for _rt in _resolved_tiers)
                            _stacked.append(_path)
                            
                        try:
                            df = pd.DataFrame(df)[_cols]
                            df.columns = pd.MultiIndex.from_tuples(_stacked)
                        except Exception:
                            pass

                    try:
                        st.dataframe(df, width="stretch", hide_index=True)
                    except TypeError:
                        st.dataframe(df, width="stretch")

                    if pd is not None:
                        csv_bytes = pd.DataFrame(df).to_csv(index=False).encode("utf-8")
                        st.download_button(
                            "Download table as CSV",
                            data=csv_bytes,
                            file_name=f"{code_mode.lower()}_table_{str(selected).lower()}.csv",
                            mime="text/csv",
                        )

    with qr_tab:
        if code_mode == "OESC":
            QR_DIR = CONTENT_DIR / "quick_reference"
            pdf_files = sorted(QR_DIR.glob("*.pdf")) if QR_DIR.exists() else []

            if not pdf_files:
                st.info("No quick reference PDFs found in `content/quick_reference/`.")
            else:
                def _qr_label(p) -> str:
                    return p.stem.replace("_", " ").replace("-", " — ")

                selected_qr = st.selectbox(
                    "Select a reference table",
                    pdf_files,
                    format_func=_qr_label,
                    key="qr_select",
                )

                if selected_qr:
                    pdf_bytes = selected_qr.read_bytes()

                    from streamlit_pdf_viewer import pdf_viewer
                    pdf_viewer(pdf_bytes, width=700, height=800)

                    st.download_button(
                        "⬇ Download PDF",
                        data=pdf_bytes,
                        file_name=selected_qr.name,
                        mime="application/pdf",
                    )


# ============================
# 11) Voltage Drop  (FULL BLOCK — Table D3 expander always shown; f-list filtered for DC; size order matches Table D3)
# ============================
elif page == "Voltage Drop":
    with theory_tab:
        header("Voltage Drop — Theory")
        show_code_note(code_mode)
        render_md_for_code("voltage_drop", code_mode)

        # -------------------------------------------------
        # Display the System Factor (f) lookup table
        # -------------------------------------------------
        with st.expander("📐 Show system factor (f) reference table", expanded=False):
            st.markdown("### System factor (f) — reference table (from Appendix D)")
            
            system_factor_data = [
                {"System / Connection": "DC — 2-wire (positive-to-negative)", "f (used in formula)": 2.0, "Voltage reference": "Positive-to-negative"},
                {"System / Connection": "DC — 2-wire (positive-to-ground)", "f (used in formula)": 2.0, "Voltage reference": "Positive-to-ground"},
                {"System / Connection": "DC — 2-wire (negative-to-ground)", "f (used in formula)": 2.0, "Voltage reference": "Negative-to-ground"},
                {"System / Connection": "DC — 3-wire, line-to-line with grounded conductor", "f (used in formula)": 2.0, "Voltage reference": "Line-to-line"},
                {"System / Connection": "1-φ AC — 2-wire, line-to-grounded conductor", "f (used in formula)": 2.0, "Voltage reference": "Line-to-ground"},
                {"System / Connection": "1-φ AC — 2-wire, line-to-line", "f (used in formula)": 2.0, "Voltage reference": "Line-to-line"},
                {"System / Connection": "1-φ AC — 3-wire, line-to-line, with grounded conductor", "f (used in formula)": 2.0, "Voltage reference": "Line-to-line"},
                {"System / Connection": "3-φ AC — 2-wire, line-to-grounded conductor", "f (used in formula)": 2.0, "Voltage reference": "Line-to-ground"},
                {"System / Connection": "3-φ AC — 2-wire, line-to-line, no grounded conductor", "f (used in formula)": 2.0, "Voltage reference": "Line-to-line"},
                {"System / Connection": "3-φ AC — 3-wire, line-to-line with grounded conductor", "f (used in formula)": 2.0, "Voltage reference": "Line-to-line"},
                {"System / Connection": "3-φ AC — 3-wire, line-to-grounded conductor", "f (used in formula)": 2.0, "Voltage reference": "Line-to-ground"},
                {"System / Connection": "3-φ AC — 3-wire, line-to-line, no grounded conductor", "f (used in formula)": math.sqrt(3), "Voltage reference": "Line-to-line"},
                {"System / Connection": "3-φ AC — 4-wire, line-to-line, with grounded conductor", "f (used in formula)": math.sqrt(3), "Voltage reference": "Line-to-line"},
            ]
            
            if pd is not None:
                df_f = pd.DataFrame(system_factor_data)
                st.dataframe(df_f, width="stretch", hide_index=True)
            else:
                for r in system_factor_data:
                    st.write(f"- **{r['System / Connection']}** — f = {r['f (used in formula)']} — {r['Voltage reference']}")

            st.caption(
                "Notes: The 'Voltage reference' column shows whether the VD is line-to-line or line-to-ground for that circuit type. "
                "f = √3 ≈ 1.732 for 3-phase line-to-line measurements."
            )

    with examples_tab:
        header("Voltage Drop — Examples")
        show_code_note(code_mode)
        render_md_for_code("voltage_drop_examples", code_mode)

    with calc_tab:
        if code_mode == "NEC":
            header("Voltage Drop — Calculator", "NEC Article 210.19/215.2 exact AC voltage drop and maximum distance calculation.")
            show_code_note(code_mode)
            from nec_calc.voltage_drop.ui import render_calc as render_nec_vd
            render_nec_vd()
        else:
            header("Voltage Drop Calculator", "OESC Rule 8-102 with Appendix B Table D3 k-values.")
            show_code_note(code_mode)
            from oesc_calc.voltage_drop.ui import render_calc as render_oesc_vd
            render_oesc_vd()


# ============================
# 12) Panel Schedule
# ============================
elif page == "Panel Schedule":
    with theory_tab:
        header("Panel Schedule — Setup")
        show_code_note(code_mode)
        st.markdown(
            "- Enter header and schedule data below to build a panel schedule workbook.\n"
            "- Use the **Download Empty Template** button for a clean copy with today’s date.\n"
            "- The exported file preserves the template layout, styles, and formulas."
        )

    with examples_tab:
        header("Panel Schedule — Tips")
        show_code_note(code_mode)
        st.markdown(
            "- Circuit numbers and phase grouping follow the template row order (A/B/C in repeating rows).\n"
            "- Use consistent units for connected load (W).\n"
            "- Leave fields blank if not applicable; totals will update automatically."
        )

    with calc_tab:
        header("Panel Schedule Builder", "Fill out fields and export to a matching Excel panel schedule.")
        show_code_note(code_mode)

        if not PANEL_TEMPLATE_PATH.exists():
            st.error(f"Panel schedule template not found: {PANEL_TEMPLATE_PATH}")
            st.stop()

        today = datetime.today().date()

        st.markdown("### Cover Sheet")
        c1, c2, c3 = st.columns([1, 1, 1], gap="large")
        with c1:
            client = st.text_input("Client / Company", value="", key="ps_client")
            facility = st.text_input("Facility / Project", value="", key="ps_facility")
            building = st.text_input("Building / Area", value="", key="ps_building")
            panel_tag = st.text_input("Panel Tag", value="", key="ps_panel_tag")
        with c2:
            title = st.text_input("Drawing Title", value="", key="ps_title")
            cover_building = st.text_input("Cover Building Title", value="", key="ps_cover_building")
            cover_title = st.text_input("Cover Schedule Title", value="", key="ps_cover_title")
            drawing_number = st.text_input("Drawing Number", value="", key="ps_drawing_number")
        with c3:
            project_number = st.text_input("Project Number", value="", key="ps_project_number")
            drawn_by = st.text_input("Drawn By", value="", key="ps_drawn_by")
            checked_by = st.text_input("Checked By", value="", key="ps_checked_by")
            approved_by = st.text_input("Approved By", value="", key="ps_approved_by")

        d1, d2, d3 = st.columns([1, 1, 1], gap="large")
        with d1:
            drawn_date = st.date_input("Drawn Date", value=today, key="ps_drawn_date")
        with d2:
            checked_date = st.date_input("Checked Date", value=today, key="ps_checked_date")
        with d3:
            approved_date = st.date_input("Approved Date", value=today, key="ps_approved_date")

        notes = st.text_area("Notes / References (Cover)", value="", height=160, key="ps_notes")

        st.markdown("### Revision Block")
        r1, r2, r3, r4 = st.columns([1, 1, 1, 1], gap="large")
        with r1:
            rev_no = st.text_input("Rev No.", value="", key="ps_rev_no")
            revised_by = st.text_input("Revised By", value="", key="ps_revised_by")
        with r2:
            rev_date = st.date_input("Revision Date", value=today, key="ps_rev_date")
            rev_checked_by = st.text_input("Revision Checked By", value="", key="ps_rev_checked_by")
        with r3:
            rev_checked_date = st.date_input("Revision Checked Date", value=today, key="ps_rev_checked_date")
            rev_approved_by = st.text_input("Revision Approved By", value="", key="ps_rev_approved_by")
        with r4:
            rev_approved_date = st.date_input("Revision Approved Date", value=today, key="ps_rev_approved_date")
            rev_notes = st.text_input("Revision Notes", value="", key="ps_rev_notes")

        st.divider()
        st.markdown("### Panel Header")
        h1, h2, h3 = st.columns([1, 1, 1], gap="large")
        with h1:
            distribution_board_no = st.text_input("Distribution Board No.", value="", key="ps_distribution_board_no")
            bus_rating = st.text_input("Bus Rating", value="", key="ps_bus_rating")
            num_circuits = st.number_input("Number of Circuits", min_value=1, value=24, step=1, key="ps_num_circuits")
            incomer_transformer_tag = st.text_input("Incomer Transformer Tag No.", value="", key="ps_incomer_transformer_tag")
        with h2:
            location = st.text_input("Location", value="", key="ps_location")
            incomer_transformer_rating = st.text_input("Incomer Transformer Rating", value="", key="ps_incomer_transformer_rating")
            incomer_breaker = st.text_input("Incomer Breaker", value="", key="ps_incomer_breaker")
            interrupting_capacity = st.text_input("Interrupting Capacity", value="", key="ps_interrupting_capacity")
        with h3:
            document_number = st.text_input("Document Number", value="", key="ps_document_number")
            equipment_layout_number = st.text_input("Equipment Layout Number", value="", key="ps_equipment_layout_number")

        st.divider()
        st.markdown("### Schedule Entries")
        if "panel_left_rows" not in st.session_state:
            st.session_state["panel_left_rows"] = default_panel_left_rows()
        if "panel_right_rows" not in st.session_state:
            st.session_state["panel_right_rows"] = default_panel_right_rows()

        lcol, rcol = st.columns([1, 1], gap="large")
        with lcol:
            st.markdown("#### Left Side (Odd Circuits)")
            left_rows = st.data_editor(
                st.session_state["panel_left_rows"],
                key="panel_left_editor",
                num_rows="fixed",
                hide_index=True,
                column_config={
                    "Cct No": st.column_config.NumberColumn("Cct No", width="small"),
                    "Phase": st.column_config.TextColumn("Phase", width="small"),
                    "Load Description": st.column_config.TextColumn("Load Description", width="medium"),
                    "Conn Load (W)": st.column_config.TextColumn("Conn Load (W)", width="small"),
                    "RCCB Rating": st.column_config.TextColumn("RCCB Rating", width="small"),
                    "No. of Fixt.": st.column_config.TextColumn("No. of Fixt.", width="small"),
                    "Brkr Size": st.column_config.TextColumn("Brkr Size", width="small"),
                },
                disabled=["Cct No", "Phase"],
            )
            st.session_state["panel_left_rows"] = left_rows
        with rcol:
            st.markdown("#### Right Side (Even Circuits)")
            right_rows = st.data_editor(
                st.session_state["panel_right_rows"],
                key="panel_right_editor",
                num_rows="fixed",
                hide_index=True,
                column_config={
                    "Cct No": st.column_config.NumberColumn("Cct No", width="small"),
                    "Brkr Size": st.column_config.TextColumn("Brkr Size", width="small"),
                    "No. of Fixt.": st.column_config.TextColumn("No. of Fixt.", width="small"),
                    "RCCB Rating": st.column_config.TextColumn("RCCB Rating", width="small"),
                    "Conn Load (W)": st.column_config.TextColumn("Conn Load (W)", width="small"),
                    "Load Description": st.column_config.TextColumn("Load Description", width="medium"),
                },
                disabled=["Cct No"],
            )
            st.session_state["panel_right_rows"] = right_rows

        st.divider()
        st.markdown("### Exports")
        exp1, exp2 = st.columns([1, 1], gap="large")

        cover_data = {
            "client": client,
            "facility": facility,
            "building": building,
            "title": title,
            "panel_tag": panel_tag,
            "drawn_by": drawn_by,
            "checked_by": checked_by,
            "approved_by": approved_by,
            "drawn_date": drawn_date,
            "checked_date": checked_date,
            "approved_date": approved_date,
            "project_number": project_number,
            "drawing_number": drawing_number,
            "cover_building": cover_building,
            "cover_title": cover_title,
            "notes": notes,
            "rev_no": rev_no,
            "revised_by": revised_by,
            "rev_date": rev_date,
            "rev_checked_by": rev_checked_by,
            "rev_checked_date": rev_checked_date,
            "rev_approved_by": rev_approved_by,
            "rev_approved_date": rev_approved_date,
            "rev_notes": rev_notes,
        }

        panel_header = {
            "distribution_board_no": distribution_board_no,
            "location": location,
            "document_number": document_number,
            "bus_rating": bus_rating,
            "incomer_transformer_rating": incomer_transformer_rating,
            "equipment_layout_number": equipment_layout_number,
            "num_circuits": int(num_circuits) if num_circuits else None,
            "incomer_breaker": incomer_breaker,
            "incomer_transformer_tag": incomer_transformer_tag,
            "interrupting_capacity": interrupting_capacity,
        }

        with exp1:
            if st.button("Prepare Filled Excel (.xlsx)", key="ps_build_filled"):
                try:
                    wb = build_panel_schedule_workbook(
                        cover_data,
                        panel_header,
                        left_rows,
                        right_rows,
                        use_current_date=False,
                    )
                    buf = io.BytesIO()
                    wb.save(buf)
                    st.session_state["ps_filled_bytes"] = buf.getvalue()
                    st.success("Filled Excel prepared. Use the download button below.")
                except Exception as e:
                    st.error(f"Failed to build panel schedule: {e}")

            filled_bytes = st.session_state.get("ps_filled_bytes", None)
            st.download_button(
                "Download Filled Excel",
                data=filled_bytes if filled_bytes else b"",
                file_name="Panel_Schedule_Filled.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                disabled=(filled_bytes is None),
                key="ps_download_filled",
            )

        with exp2:
            if st.button("Prepare Empty Template (.xlsx)", key="ps_build_template"):
                try:
                    wb = build_panel_schedule_workbook(
                        {
                            "client": "",
                            "facility": "",
                            "building": "",
                            "title": "",
                            "panel_tag": "",
                            "drawn_by": "",
                            "checked_by": "",
                            "approved_by": "",
                            "drawn_date": today,
                            "checked_date": today,
                            "approved_date": today,
                            "project_number": "",
                            "drawing_number": "",
                            "cover_building": "",
                            "cover_title": "",
                            "notes": "",
                            "rev_no": "",
                            "revised_by": "",
                            "rev_date": today,
                            "rev_checked_by": "",
                            "rev_checked_date": today,
                            "rev_approved_by": "",
                            "rev_approved_date": today,
                            "rev_notes": "",
                        },
                        {
                            "distribution_board_no": "",
                            "location": "",
                            "document_number": "",
                            "bus_rating": "",
                            "incomer_transformer_rating": "",
                            "equipment_layout_number": "",
                            "num_circuits": 24,
                            "incomer_breaker": "",
                            "incomer_transformer_tag": "",
                            "interrupting_capacity": "",
                        },
                        default_panel_left_rows(),
                        default_panel_right_rows(),
                        use_current_date=True,
                    )
                    buf = io.BytesIO()
                    wb.save(buf)
                    st.session_state["ps_template_bytes"] = buf.getvalue()
                    st.success("Empty template prepared. Use the download button below.")
                except Exception as e:
                    st.error(f"Failed to build template: {e}")

            template_bytes = st.session_state.get("ps_template_bytes", None)
            st.download_button(
                "Download Empty Template",
                data=template_bytes if template_bytes else b"",
                file_name="Panel_Schedule_Template.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                disabled=(template_bytes is None),
                key="ps_download_template",
            )


# ============================
# 13) Conductors
# ============================
elif page == "Conductors":
    with theory_tab:
        header("Conductors — Theory")
        show_code_note(code_mode)
        render_md_for_code("conductors", code_mode)

    with examples_tab:
        header("Conductors — Examples")
        show_code_note(code_mode)
        render_md_for_code("conductors_examples", code_mode)

    with calc_tab:
        if code_mode == "NEC":
            header("Conductors — Calculator", "NEC Article 310 ampacity verification, ambient temperature correction, adjustment factors, and terminal temperature limits.")
            show_code_note(code_mode)
            from nec_calc.conductors.ui import render_calc as render_nec_cond
            render_nec_cond()
        else:
            header("Conductors — Calculator", "OESC Rule 4-004 ampacity table selection and correction factors.")
            show_code_note(code_mode)
            from oesc_calc.conductors.ui import render_calc as render_oesc_cond
            render_oesc_cond()
